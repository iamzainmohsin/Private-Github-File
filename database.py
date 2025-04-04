import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os

class Task():
    def __init__(self, description, due_date, status, index = 0):
        self.index = index
        self.description = description
        self.due_date = self.format_date(due_date)
        self.status = status

    @staticmethod
    def format_date(date_str):
        try:
            # Attempt to parse the date in various common formats
            date_formats = ["%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%B-%Y"]
            for fmt in date_formats:
                try:
                    parsed_date = datetime.strptime(date_str, fmt)
                    return parsed_date.strftime("%d-%B-%Y")  # Output in date-Month-Year format
                except ValueError:
                    continue
            raise ValueError("Invalid date format")  # If no format matches
        except Exception as e:
            return f"Error formatting date: {e}"


    # Output!
    def __str__(self):
        # User-friendly representation
        return f"{self.index}.\nTask: '{self.description}'\nDue Date: {self.due_date}\nCurrent status: {self.status}\n"


class ExcelHandler():
    
    def __init__(self, file_name):
        self.file_name = file_name
        self.workbook = Workbook()


    #Set Headers:
    def set_headers(self, worksheet1, worksheet2):
        header_text = ["Task Description", "Due Date", "Status"]
        worksheet1.append(header_text)
        worksheet2.append(header_text)
        print(f"Created Sheets {worksheet1.title} and {worksheet2.title}")


    #sets up the Workbook!
    def workbook_setup(self):
        if os.path.exists(self.file_name):
            self.workbook = openpyxl.load_workbook(self.file_name)
            print(f"Workbook has been loaded!")

            #checks for headers in Already existing sheet:
            for sheet_name in ["Pending", "Completed"]:
                if sheet_name in self.workbook.sheetnames:
                    workheet = self.workbook[sheet_name]

        else:    
            del self.workbook['Sheet']
            worksheet_pending = self.workbook.create_sheet(title="Pending")
            worksheet_completed = self.workbook.create_sheet(title="Completed")

            #headers Setup:
            self.set_headers(worksheet_pending, worksheet_completed)
            self.workbook.save(self.file_name)
            print(f"New Workbook has been created!")


    #Writes to the Workbook!  
    def add_tasks(self, task: Task):
        worksheet = self.workbook["Pending"]
        row_data = [task.description, task.due_date, task.status]
        worksheet.append(row_data)
        self.workbook.save(self.file_name)
        print(f"DEBUG: Adding task - {task.description}, {task.due_date}, {task.status}")  # Debugging line
        return True


    #Reading from the Workbook!  
    def get_tasks(self):
        tasks_list = []
        worksheet = self.workbook["Pending"]
        
        #Reads the Data:
        for index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start = 1):
            try:
                if row and len(row)>=3:
                    description, due_date, status = row
                    print(f"Task ID: {index}, Description: {description}")
                    task_obj = Task(index=index, description=description, due_date=due_date, status=status)
                    tasks_list.append(task_obj)
                    print(task_obj)
                else:
                    print(f"Skipping incomplete row: {row}")
            except Exception as e: 
                print(f"Error processing row {index}: {e}")
                continue       


        return tasks_list    


    #Updating Status of the Task  
    def complete_task(self, task_id, new_details):
        worksheet = self.workbook["Pending"]
        task_id = int(task_id)

        #Edge Cases fix:
        valid_statuses = ["P", "C"]
        if new_details not in valid_statuses:
            print(f"Invalid status '{new_details}'. Use 'C' for Completed or 'P' for Pending.")
            return

        for index, row in enumerate(worksheet.iter_rows(min_row=2), start = 1):
            if index == task_id:
                row[2].value = new_details      #Row2 == Status column
                self.workbook.save(self.file_name)
                print(f"{task_id} marked as Completed")
                return

        
    #Moving Completed Tasks  
    def move_complete_task(self):
        pending_worksheet = self.workbook["Pending"]
        completed_worksheet = self.workbook["Completed"]
        rows_to_delete = []

        #moving to Completed Sheet:
        for row_index, row in enumerate(pending_worksheet.iter_rows(min_row=2), start = 2):
                if row[2].value == "C":      #Row2 == Status column
                    completed_worksheet.append([row[0].value, row[1].value, row[2].value])
                    rows_to_delete.append(row_index)
        
        #Edge Case
        if not rows_to_delete:
            print("No completed tasks to move.")
            return

        #properly deleting Completed Rows:
        for row_index in reversed(rows_to_delete):
            pending_worksheet.delete_rows(row_index)
            print(f"Row Deleted: {rows_to_delete}")

        self.workbook.save(self.file_name)
        print("Completed tasks have been moved to the 'Completed' sheet.")


    #Deleting Tasks:
    def delete_task(self, task_id):
        # Validate the task_id
        if not task_id.isdigit():
            print(f"Invalid Task ID: '{task_id}'. Please choose Numeric Values only")
            return False

        worksheet = self.workbook["Pending"]
        print(f"Searching for Task ID '{task_id}' in sheet: {worksheet}")

        # Start enumerating from 1 while manually skipping the header
        task_deleted = False
        for row_index, row in enumerate(worksheet.iter_rows(min_row=1), start=1):
            if row_index == 1:  # Skip the header row explicitly
                continue

            # Check if task_id matches the row index (adjust for Python's indexing)
            if str(row_index - 1) == str(task_id):  # Subtract 1 to align with task numbering logic
                worksheet.delete_rows(row_index)  # Delete the row
                self.workbook.save(self.file_name)  # Save the updated workbook
                print(f"Task ID '{task_id}' successfully deleted from sheet: {worksheet}.")
                task_deleted = True
                break

        print(f"Task ID '{task_id}' not found in sheet: {worksheet}.")
        return task_deleted


   #Summary:
    def generate_report(self):
        pending_worksheet = self.workbook["Pending"]
        completed_worksheet = self.workbook["Completed"]

        #Counters:
        total_tasks = 0
        pending_count = 0
        completed_count = 0
        overdue_count = 0
        overdue_tasks = []
        today = datetime.today()

        #Counts Pending
        for row in pending_worksheet.iter_rows(min_row=2, values_only=True):
            total_tasks += 1
            if row[2] == "P":
                pending_count += 1

                formatted_date = Task.format_date(row[1])
                due_date = datetime.strptime(formatted_date, "%d-%B-%Y")

                if due_date < today:
                    overdue_count += 1
                    overdue_tasks.append(row)



        #Counts Completed
        for row in completed_worksheet.iter_rows(min_row=2, values_only=True):
            total_tasks += 1
            completed_count += 1   

        #Report:
        summary = f"**Total Tasks:** {total_tasks}\n**Pending Tasks:** {pending_count}\n**Overdue Tasks:** {overdue_count}\n**Completed Tasks:** {completed_count}"       
        print(summary)

        # Optional: Print overdue tasks
        if overdue_tasks:
            print("Overdue Tasks:")
            for task in overdue_tasks:
                print(f"Description: {task[0]}, Due Date: {task[1]}")

        return summary